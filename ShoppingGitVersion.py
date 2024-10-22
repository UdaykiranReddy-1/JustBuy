from tkinter import *
from tkinter.font import BOLD
import tkinter.messagebox as tsmg
from PIL import ImageTk, Image
from openpyxl import load_workbook
from tkinter import ttk
from datetime import *
import requests
import random
import json
import pandas as pd


root=Tk()
root.withdraw()
root.title('PE$U Super-M@rt')
width=root.winfo_screenwidth()
height=root.winfo_screenheight()
root.geometry("%dx%d"%(width,height))
first=Toplevel(root)
first.geometry("600x550")
first.title("OTP Verification")
second=Toplevel(root)
second.withdraw()
second.geometry("800x450")
second.title("My Cart")
third=Toplevel(root)	
third.geometry("760x430")
third.title("*$BILL$*")
third.withdraw()		

now=datetime.now()
d=now.strftime("%d/%m/%Y")
t=now.strftime("%H:%M:%S")

def on_closing():
	if tsmg.askokcancel("Quit", "Do you want to quit?"):
		first.destroy()
		root.destroy()
		second.destroy()
first.protocol("WM_DELETE_WINDOW", on_closing)		



#+++++++++++=========++++++++======= OTP SECTION =======++++++++++++============++++++++++
OTP=random.randint(1,999999)
print(OTP)
def sms_send():
    
    url = "https://www.fast2sms.com/dev/*****"  # Put your URL here 

    my_data = {
	# Your default Sender ID
	'sender_id': '*****',   # put your senderID here
	
	# Put your message here!
	'message': 'This is a test message',
	'language': 'english',
	'route': 'p',
	
	# You can send sms to multiple numbers
	# separated by comma.
	'numbers': ''
    }
    my_data['numbers']=a
    my_data['message']=f"Hello {n},\nYour One Time Password(OTP) for verification and login is {OTP}\nWish you a Happy Shopping :-)\n-PE$U SuperM@rt"
    # create a dictionary
    headers = {
        'authorization': '*********************************************************', # you need to fill in your authorization headers
        'Content-Type': "application/x-www-form-urlencoded",
        'Cache-Control': "no-cache"
    }
    # make a post request
    response = requests.request("POST",
                                url,
                                data = my_data,
                                headers = headers)

    #load json data from source
    returned_msg = json.loads(response.text)
        
def send():
	global a,n,address
	a=num.get()
	n=name.get()
	address=ad.get()
	if(a==""):
		tsmg.showerror("Error","Enter Your Mobile Number")
	elif (len(a)<10):
		tsmg.showerror("Error","Invalid Mobile Number")
		num.set("")
	else:
		b=tsmg.askyesno("Info",f"Your Number is {a} ?")
		if(b==True):
			sms_send()
		else:
			num.set("")

def check():
	c=otp.get()
	if(c==""):
		tsmg.showerror("Error","Enter the OTP")
	else:
		if(str(OTP)==c):
			tsmg.showinfo("Info","Successfully verified,Login succesful")
			
			first.destroy()
			root.deiconify()
			
		else:
			tsmg.showerror("Error","Invalid OTP")
			first.destroy()
			
num=StringVar()
otp=StringVar()
name=StringVar()
ad=StringVar()

f1=Frame(first)
Label(f1,text="OTP Verification",font="SegoeUI 30 bold",fg="purple").pack(padx=5,pady=10)
f1.pack(fill=BOTH)

f2=Frame(first)
Label(f2,text="Enter your Name: ",font="SegoeUI 20 bold",fg="teal").pack(padx=5,pady=5)
e1=Entry(f2,textvariable=name,font="SegoeUI 14 bold",fg="black",bg="white",relief=SUNKEN,borderwidth=4,justify="center").pack(ipady=5)
Label(f2,text="Enter Your Number: ",font="SegoeUI 20 bold",fg="teal").pack(padx=5,pady=5)
e2=Entry(f2,textvariable=num,font="SegoeUI 14 bold",fg="black",bg="white",relief=SUNKEN,borderwidth=4,justify="center").pack(ipady=5)
Label(f2,text="Enter Your Place: ",font="SegoeUI 20 bold",fg="teal").pack(padx=5,pady=5)
e3=Entry(f2,textvariable=ad,font="SegoeUI 14 bold",fg="black",bg="white",relief=SUNKEN,borderwidth=4,justify="center").pack(ipady=5)
f2.pack(fill=BOTH,padx=5,pady=10)

f3=Frame(first)
Label(f3,text="Enter OTP",font="SegoeUI 20 bold",fg="teal").pack(padx=5,pady=5)
e2=Entry(f3,textvariable=otp,font="SegoeUI 14 bold",fg="black",bg="white",relief=SUNKEN,borderwidth=5,justify="center").pack(ipady=5)
f3.pack(fill=BOTH,padx=5,pady=10)

f4=Frame(first)
Button(f4,text="Send OTP",command=send,font="SegoeUI 10 bold",fg="purple").pack(padx=20,pady=10,side=LEFT)
Button(f4,text="Verify OTP",command=check,font="SegoeUI 10 bold",fg="purple").pack(padx=40,pady=10,side=LEFT)
f4.pack()
#==========================================OTP SECTION END===================================================

#==========================================MAIN WINDOW FORMATTING====================================================

# Define image
bg = ImageTk.PhotoImage(file="sprmktfinal.png")

# Create a canvas
my_canvas = Canvas(root,  width=1366, height=768)
my_canvas.pack(fill="both", expand=True)

# Set image in canvas
my_canvas.create_image(0,0, image=bg, anchor="nw")

# Add a label
my_canvas.create_text(650, 30, text="Welcome to PE$U-SuperM@rt!", font=("Helvetica", 40), fill="white")
my_canvas.create_text(200, 180, text="Select the Category:- ...", font=("Helvetica", 20), fill="white")
my_canvas.create_text(590, 180, text="Select the Item:- ...", font=("Helvetica", 20), fill="white")
my_canvas.create_text(1045, 180, text="Enter number of units:- ...", font=("Helvetica", 20), fill="white")
my_canvas.create_text(1000, 270, text="Item Selected:-...", font=("Helvetica", 20), fill="white")
my_canvas.create_text(1000, 364, text="Price of Item :-...", font=("Helvetica", 20), fill="white")

# adding entries
nunits=Entry(root,font=("Helvetica",16),width=22,fg="black",justify=CENTER,bd=0)
my_canvas.create_window(1100,220,window=nunits)

entereditem=Entry(root,font=("Helvetica",16),width=22,fg="black",justify=CENTER,bd=0)
my_canvas.create_window(1100,310,window=entereditem)

entereditemprice=Entry(root,font=("Helvetica",16),width=22,fg="black",justify=CENTER,bd=0)
my_canvas.create_window(1100,400,window=entereditemprice)

#===============================MENU BAR BEGINS===============================================================
def selecting():
	tsmg.showinfo("Selecting item","To select a item, first select the category of item you want to buy in the left list box\nThen items available in that category will be displayed on right listbox\nSelect the item you want to buy")
def knowprice():
	tsmg.showinfo("Knowing Price of an Item","The price of the item you want to buy is displayed just next to the item name separated by colon\nThe numbers you see next to the name of the item is the price of the one unit in standard units of purchase i.e., 1kg for grains,flours,fruits,vegetables,etc, and 1unit for soaps,biscuit packs,etc,.. \nSo this is how the price is displayed ")
def additem():
	tsmg.showinfo("Add items","To add an item to the cart, enter the quantity of the items you want to buy and then click the 'Add to Cart' button")
def viewcart():
	tsmg.showinfo("View items","To view your selected items click on the 'View My Cart' cart, it will take you to a new window where you can view and modify the selected items")
def bill():
	tsmg.showinfo("Display Bill","After you are done shopping, click the button to generate the bill, this can also be done in the view my cart window")
def about():
	tsmg.showinfo("About us","Pesu-SuperMart is a program designed to cater the needs of the customers who, due to various reasons, are unable to go out and want an easy way to shop things which are necessary in everyday life")
def contact():
	tsmg.showinfo("Contact us","email: pesusupermart2022@gmail.com\nPhone: 4865412345")
def continue1():
	tsmg.showinfo("Continue Shopping","If you still want to continue shopping feel free to click the 'Continue Shopping' button")
def quit1():
	tsmg.showinfo("Quit shopping","If you are not happy with our services, feel free to click the 'Quit' button to discontinue shopping and exit the program")
def update1():
	tsmg.showinfo("Update quantity","If you change your mind and want to shop more items of a selected product, click on the required item and in the entry box enter the number of units you need and click the 'Update' button")
def bill1():
	tsmg.showinfo("Generate Bill","If you are done shopping click on this button to display the total amount of the selected items")
	# Creating Menubar
menubar = Menu(root)
  
# Adding File Menu and commands
Mainpage_Instructions = Menu(menubar, tearoff = 0)
menubar.add_cascade(label ='Mainpage Instructions', menu = Mainpage_Instructions)
Mainpage_Instructions.add_command(label ='Selecting Item', command =selecting )
Mainpage_Instructions.add_command(label ='Knowing Price of Item', command =knowprice )
Mainpage_Instructions.add_command(label ='Adding Item to Cart', command =additem )
Mainpage_Instructions.add_command(label ='View Cart', command =viewcart )
Mainpage_Instructions.add_command(label ='Generate Bill', command = bill)

# Adding Cart Menu and commands
Cart_instructions = Menu(menubar, tearoff = 0)
menubar.add_cascade(label ='Cart Instructions', menu = Cart_instructions)
Cart_instructions.add_command(label ='Update Quantity', command = update1)
Cart_instructions.add_command(label ='Continue Shopping', command = continue1)
Cart_instructions.add_command(label ='Generate Bill', command = bill1)
Cart_instructions.add_command(label ='Quit', command = quit1)

#Adding about us menu and commands
About_us = Menu(menubar, tearoff = 0)
menubar.add_cascade(label ='About us', menu = About_us)
About_us.add_command(label ='Know about Us', command = about)
About_us.add_separator()
About_us.add_command(label ='Contact', command = contact)
root.config(menu=menubar)
#===================================MENU BAR ENDS====================================================================
#===================================LIST BOX START===================================================================
my_listbox=Listbox(root,width=30,height=20)
my_listbox.place(x=130,y=200)
my_listbox.configure(bg='black',fg='white',font=('Aerial 13'))

wb=load_workbook('Book1.xlsx')
ws=wb.active

col_a = ws["A"]
col_b = ws["B"]
col_c = ws["C"]
col_d = ws["D"]
col_e = ws["E"]
col_f = ws["F"]
col_g = ws["G"]
col_h = ws["H"]
col_i = ws["I"]

for items in col_a:
    my_listbox.insert(END,items.value)

my_listbox2=Listbox(root,width=30,height=20)
my_listbox2.place(x=550,y=200)
my_listbox2.configure(bg='black',fg='white',font=('Aerial 13'))

Grains=col_b
Books=col_c
Fruits=col_d
soaps=col_e
biscuits=col_f
cosmetics=col_g
flours=col_h
dryfruits=col_i

#===================================LIST BOX END========================================================================

#===================================FUNCTIONS FOR MAIN PAGE=======================================================================
def list_color(e):
	my_listbox2.delete(0, END)
	if my_listbox.get(ANCHOR) == "Grains & cereals & spices":
		for item in Grains:
			my_listbox2.insert(END, item.value)
	if my_listbox.get(ANCHOR) == "Stationary & Books":
		for item in Books:
			my_listbox2.insert(END, item.value)
	if my_listbox.get(ANCHOR) == "fruits & Vegetables":
		for item in Fruits:
			my_listbox2.insert(END, item.value)
	if my_listbox.get(ANCHOR) == "Soaps & Detergents":
		for item in soaps:
			my_listbox2.insert(END, item.value)
	if my_listbox.get(ANCHOR) == "Flours":
		for item in flours:
			my_listbox2.insert(END,item.value)
	if my_listbox.get(ANCHOR) == "Dryfruits":
		for item in dryfruits:
			my_listbox2.insert(END,item.value)
	
	if my_listbox.get(ANCHOR) == "Cosmetics":
		for item in cosmetics:
			my_listbox2.insert(END, item.value)
	if my_listbox.get(ANCHOR) == "Biscuts & Foods":
		for item in biscuits:
			my_listbox2.insert(END, item.value)
# Bind The Listbox
my_listbox.bind("<<ListboxSelect>>", list_color)
wb.save("Book1.xlsx")

def itemselected_(event):
	global price,item
	selection = event.widget.curselection()
	if selection:
		nunits.delete(0,END)
		entereditem.delete(0,END)
		entereditemprice.delete(0,END)
		index=selection[0]
		data=event.widget.get(index)
		d=str(data)
		item,price= d.split(":")
		entereditem.insert(0,item)
		entereditemprice.insert(0,price)
my_listbox2.bind('<<ListboxSelect>>',itemselected_)

def add_item():
	req=nunits.get()
	if req != "" :
		nunits.delete(0,END)
		entereditem.delete(0,END)
		entereditemprice.delete(0,END)
		wb = load_workbook("itemt.xlsx")
		ws = wb.active
		ws['A1']="Item"             #category selected also could be put if if tried using binding
		ws['B1']="Unit_price" 
		ws['C1']="Number_of_units" 
		ws['D1']="Total" 
		# ws.append([item,price,req,Total])
		ws['A2']= item
		ws['B2']= int(price)
		ws['C2']= int(req)
		ws['D2']=((ws['B2'].value)*(ws['C2'].value))
		wb.save("itemt.xlsx")
		file1 = pd.read_excel("itemp.xlsx")
		file2 = pd.read_excel("itemt.xlsx")
		all = [file1,file2]
		append= pd.concat(all)
		append.to_excel("itemp.xlsx",index= False )
		fillagain()
		
#==========================FUNCTIONS OF CART=================================================================================
def destroycart():
	second.withdraw()

def destroy_whole():
	x=tsmg.askyesno('*ALERT*',"Caution!! This would empty your cart and delete all data\nAre you sure to exit?")
	if x==True:
		second.destroy()
		root.destroy()
	else :
		second.withdraw()

def clear_tree():
	global my_tree
	my_tree.delete(*my_tree.get_children())	#clear old treeview 	

def view_cart():
	second.deiconify()

# def deleteitem():
# 	a=my_tree.index(my_tree.selection())
# 	df4=pd.read_excel("itemp.xlsx")
# 	p=df4.iat[int(a),0]
# 	df4.drop(p,axis=0,inplace=True)
# 	fillagain()
	
def select_record(e):
    #clear entry boxes
    itemsecond.set("")
    itemprice.set("")
    itemunits.set("")
    #grab record number
    selected=my_tree.focus()
    #grab record values
    values =my_tree.item(selected,'values')
    #outputs to entryboxes
    itemsecond.set(f"{values[0]}")
    itemprice.set(f"{values[1]}")
    itemunits.set(f"{values[2]}")

def update():
	if itemunits.get()!= "":
		
		a=my_tree.index(my_tree.selection())
		df1=pd.read_excel("itemp.xlsx")
		df1.iat[a,2]=itemunits.get()
		df1.iat[a,3]= ((df1.iat[a,2])*(df1.iat[a,1]))
		df1.to_excel("itemp.xlsx",index=False)
		itemsecond.set("")
		itemprice.set("")
		itemunits.set("")
		fillagain()
		
def fillagain():
	# Clear the Treeview
	for record in my_tree.get_children():
		my_tree.delete(record)
	df2=pd.read_excel("itemp.xlsx")

	my_tree["column"]=list(df2.columns)	#set up new treeview
	my_tree["show"]="headings"
	for column in my_tree["column"]:
		my_tree.heading(column,text=column)

	df2_rows=df2.to_numpy().tolist()
	for row in df2_rows:
		my_tree.insert("","end",values=row)

def gen_bill():
	global z
	x=tsmg.askyesno("NOTE","This is the last step of shopping\nIf you want to view the items selected and update quantity,go to view my cart\nOr you want to end shopping and pay bill ,continue with yes..")
	if x==1:
		second.destroy()
		third.deiconify()
		
		df = pd.read_excel('itemp.xlsx')
		z=(df["Total"].sum()) 
		
		tree_frame2=Frame(third)
		tree_frame2.pack(pady=10)
		my_bill=ttk.Treeview(tree_frame2)
		tree_frame3=Frame(third)
		tree_frame3.pack(pady=10)
		#=============ADD SOME STYLE============================
		style2=ttk.Style()
		#=============PICK A THEME===============================
		style2.theme_use('default')
		#=============CONFIGURE THE TREEVIEW COLORS==============
		style2.configure('Treeview',background="#D3D3D3",foreground="black",rowheight=25,fieldbg="#D3D3D3")
		#=============CHANGE SELECTED COLOR=====================
		style2.map('Treeview',background=[('selected','#347083')])
		#=============CREATE STRIPED ROW TAGS======================
		my_bill.tag_configure('oddrow',background="white")
		my_bill.tag_configure('evenrow',background="lightblue")
		# #=============COLUMN WIDTHS============================
		# my_bill.column('#0',minwidth=0,width=300,stretch=NO)
		# my_bill.column('#1',minwidth=0,width=150,stretch=NO)
		# my_bill.column('#2',minwidth=0,width=150,stretch=NO)
		# my_bill.column('#3',minwidth=0,width=100,stretch=NO)

		#=============CREATE A TREEVIEW SCROLLBAR==============
		tree_scroll1=Scrollbar(tree_frame2)
		tree_scroll1.pack(side=RIGHT,fill=Y)
		#=============CREATE THE TREEVIEW=======================
		my_bill=ttk.Treeview(tree_frame2,yscrollcommand=tree_scroll1.set,selectmode="extended")
		#=============CONFIGURE SCROLLBAR==========================
		tree_scroll1.config(command=my_bill.yview)
		df=pd.read_excel("itemp.xlsx")
		my_bill["column"]=list(df.columns)	#set up new treeview
		my_bill["show"]="headings"
		for column in my_bill["column"]:
			my_bill.heading(column,text=column)
		df_rows=df.to_numpy().tolist()
		i=0
		for row in df_rows:
			if i % 2 == 0 :
				my_bill.insert("","end",values=row,tags=('evenrow',))
			else:
				my_bill.insert("","end",values=row,tags=('oddrow',))
			i+=1
		my_bill.pack()
		nettotal=Label(tree_frame3,text=f"Net Total: {z}",fg="black",font=("Helvetica",15))
		nettotal.pack(padx=10,pady=10)

		Button10=Button(tree_frame3,text="Finish",command=update_amount_xl,bd=0,fg="white",bg='orange',font=("Helvetica",15))
		Button10.pack(padx=10,pady=10)

def on_exiting():
	update_amount_xl()
third.protocol("WM_DELETE_WINDOW",on_exiting)	

def close_cart():
	second.withdraw()
second.protocol("WM_DELETE_WINDOW", close_cart)

def update_amount_xl():
	wb = load_workbook("custinfot.xlsx")     
	ws = wb.active
	ws['A1']="Name" 
	ws['B1']="Address" 
	ws['C1']="Phone Number" 
	ws['D1']="date of shopping" 
	ws['E1']="time of shopping" 
	ws['F1']="Total bill amount"
	ws['A2']= n
	ws['B2']= address 
	ws['C2']= num.get()
	ws['D2']= d
	ws['E2']= t
	if z!=0:
		ws['F2']= z
	if z==0:
		ws['F2']= "purchased none"
	wb.save("custinfot.xlsx")
	file1 = pd.read_excel("custinfop.xlsx")
	file2 = pd.read_excel("custinfot.xlsx")
	all = [file1,file2]
	append= pd.concat(all)
	append.to_excel("custinfop.xlsx",index= False )
	tsmg.showinfo("Greetings!",f"Thank you so much for shopping with us {name.get()}\nHope you will love our products and become our regular customer\nSee you back soon :)\n-PE$U SuperM@rt ")
	third.destroy()
	root.destroy()
#===============================CART FORMATTING========================================================================
	#=============CREATE A TREEVIEW FRAME======================
tree_frame=Frame(second)
tree_frame.pack(pady=10)
my_tree=ttk.Treeview(tree_frame)

#=============ADD SOME STYLE============================
style=ttk.Style()
#=============PICK A THEME===============================
style.theme_use('default')
#=============CONFIGURE THE TREEVIEW COLORS==============
style.configure('Treeview',background="#D3D3D3",foreground="black",rowheight=25,fieldbg="#D3D3D3")
#=============CHANGE SELECTED COLOR=====================
style.map('Treeview',background=[('selected','#347083')])
#=============CREATE STRIPED ROW TAGS======================
my_tree.tag_configure('oddrow',background="white")
my_tree.tag_configure('evenrow',background="lightblue")
# #=============COLUMN WIDTHS========================================
# my_tree.column('#0',minwidth=0,width=400,stretch=NO)
# my_tree.column('#1',minwidth=0,width=175,stretch=NO)
# my_tree.column('#2',minwidth=0,width=150,stretch=NO)
# my_tree.column('#3',minwidth=0,width=100,stretch=NO)

#=============CREATE A TREEVIEW SCROLLBAR==============
tree_scroll=Scrollbar(tree_frame)
tree_scroll.pack(side=RIGHT,fill=Y)
#=============CREATE THE TREEVIEW=======================
my_tree=ttk.Treeview(tree_frame,yscrollcommand=tree_scroll.set,selectmode="extended")

#=============CONFIGURE SCROLLBAR==========================
tree_scroll.config(command=my_tree.yview)
df=pd.read_excel("itemp.xlsx")

#================CREATING CART WINDOW=============================
my_tree["column"]=list(df.columns)	#set up new treeview
my_tree["show"]="headings"
for column in my_tree["column"]:
	my_tree.heading(column,text=column)

df_rows=df.to_numpy().tolist()
for row in df_rows:
	my_tree.insert("","end",values=row)
my_tree.pack()

#2 ===========BIND TREEVIEW=================================
my_tree.bind("<ButtonRelease-1>",select_record)

itemsecond=StringVar()
itemprice=IntVar()
itemunits=IntVar()
frame2=Frame(second)	
# Button10=Button(frame2,text="Delete items",command= deleteitem,bd=0,fg="black",font=("Helvetica",15)).grid(row=1,column=0,padx=10,pady=10)
Button11=Button(frame2,text="Continue Shopping",command=destroycart,bd=0,fg="white",bg='green',font=("Helvetica",15)).grid(row=1,column=1,padx=10,pady=10)
Button12=Button(frame2,text="Done Shopping!\nGenerate bill",command=gen_bill,bd=0,fg="white",bg='blue',font=("Helvetica",15)).grid(row=1,column=2,padx=10,pady=10)
Button13=Button(frame2,text="Quit Shopping",command=destroy_whole,bd=0,fg="white",bg='red',font=("Helvetica",15)).grid(row=1,column=3,padx=10,pady=10)
Button13=Button(frame2,text="Update",command=update,bd=0,fg="white",bg='purple',font=("Helvetica",15)).grid(row=1,column=4,padx=10,pady=10)
entryi=Entry(frame2,bd=0,textvariable=itemsecond,state=DISABLED,fg='black',font=("Helvetica",15),width=22,justify=CENTER).grid(row=0,column=0,columnspan=2,padx=5,pady=10)
entryp=Entry(frame2,bd=0,textvariable=itemprice,state=DISABLED,fg='black',font=("Helvetica",15),width=14,justify=CENTER).grid(row=0,column=2,columnspan=2,padx=5,pady=10)
entryn=Entry(frame2,bd=0,textvariable=itemunits,fg='black',font=("Helvetica",15),width=14,justify=CENTER).grid(row=0,column=4,columnspan=2,padx=5,pady=10)

# def on_entry_click(event):
#     """function that gets called whenever entry is clicked"""
#     if entryi.get() == 'Item Selected...':
#        entryi.delete(0, "end") # delete all the text in the entry
#        entryi.insert(0, '') #Insert blank for user input
#        entryi.config(fg = 'black')
# def on_focusout(event):
#     if entryi.get() == '':
#         entryi.insert(0, 'Item Selected...')
#         entryi.config(fg = 'grey')

															# This part is for placeholders for cart showings
# entryi.insert(0, 'Item selected...')
# entryi.bind('<FocusIn>', on_entry_click)
# entryi.bind('<FocusOut>', on_focusout)

frame2.pack()	
# add some buttons
button1 = Button(root, text="Add to Cart",command=add_item,bd=0,fg="white",bg='green',font=("Helvetica",15))
button2 = Button(root, text="View my Cart",command=view_cart,bd=0,fg="white",bg='purple',font=("Helvetica",15))
button3= Button(root,text="Generate Bill",command=gen_bill,bd=0,fg="white",bg='red',font=("Helvetica",15))

button1_window = my_canvas.create_window(1050, 430, anchor="nw", window=button1)
button2_window = my_canvas.create_window(1042, 490, anchor="nw", window=button2)
button3_window = my_canvas.create_window(1046, 550, anchor="nw", window=button3)


def resizer(e):
	global bg1, resized_bg, new_bg
	# Open our image
	bg1 = Image.open("sprmktfinal.png")
	# Resize the image
	resized_bg = bg1.resize((e.width, e.height), Image.ANTIALIAS)
	# Define our image again
	new_bg = ImageTk.PhotoImage(resized_bg)
	# Add it back to the canvas
	my_canvas.create_image(0,0, image=new_bg, anchor="nw")
	# Readd the text
	my_canvas.create_text(650, 30, text="Welcome to PE$U-SuperM@rt!", font=("Helvetica", 40), fill="white")
	my_canvas.create_text(200, 180, text="Select the Category:- ...", font=("Helvetica", 20), fill="white")
	my_canvas.create_text(590, 180, text="Select the Item:- ...", font=("Helvetica", 20), fill="white")
	my_canvas.create_text(1045, 180, text="Enter number of units:- ...", font=("Helvetica", 20), fill="white")
	my_canvas.create_text(1000, 270, text="Item Selected:-...", font=("Helvetica", 20), fill="white")
	my_canvas.create_text(1000, 364, text="Price of Item :-...", font=("Helvetica", 20), fill="white")

root.bind('<Configure>', resizer)
root.mainloop()








